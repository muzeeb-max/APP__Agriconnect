from flask import Flask, render_template, request, redirect, url_for, flash, session

from openpyxl import load_workbook, Workbook
import os
app = Flask(__name__)
app.secret_key = 'kojja_super_secret_key_123'  # 👈 can be any random string


# ---------------- EXCEL SETUP ----------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "user_data.xlsx")

def initialize_excel():
    """Create Excel file if not exists."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Full Name", "Username", "Email", "Password", "Type"])
        wb.save(EXCEL_FILE)
        print("✅ Excel file created:", EXCEL_FILE)
    else:
        print("📘 Using existing Excel file:", EXCEL_FILE)


# ---------------- ROUTES ----------------

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/contacts')
def contact():
    return render_template('contacts.html')

@app.route('/register', methods=["GET", "POST"])
def register():
    if request.method == "POST":
        fullname = request.form.get("fullname")
        username = request.form.get("username")
        email = request.form.get("email")
        password = request.form.get("password")

        if not all([fullname, username, email, password]):
            return render_template("register.html", message="⚠️ Please fill all fields")

        # Save to Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([fullname, username, email, password, ""])
        wb.save(EXCEL_FILE)

        print(f"✅ Added user: {fullname} | {username} | {email}")
        return redirect(url_for('regi', name=fullname))

    return render_template("register.html")

@app.route('/success')
def success():
    name = request.args.get('name')
    return render_template('success.html', name=name)

@app.route("/regi")
def regi():
    name = request.args.get('name')
    return render_template("reg.html", name=name)

# ---------------- SUBMIT ROLE ----------------
@app.route('/submit-role', methods=['POST'])
def submit_role():
    name = request.form.get('name')
    user_role = request.form.get('user_role')

    if not all([name, user_role]):
        return "⚠️ Missing name or role", 400

    # Load workbook
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Find matching user row and update their Type
    for row in ws.iter_rows(min_row=2):
        if row[0].value == name:
            row[4].value = user_role
            break

    wb.save(EXCEL_FILE)
    print(f"✅ Updated {name}'s role to {user_role}")

    # Redirect based on selected role
    if user_role == "farmer":
        return redirect(url_for('dashboard', name=name, Type=user_role))
    elif user_role == "vendor":
        return redirect(url_for('dashboardV', name=name, Type=user_role))
    else:
        return "⚠️ Invalid role", 400


# ---------------- LOGIN FUNCTION ----------------
@app.route('/login', methods=['POST'])
def login():
    email = request.form.get('email')
    password = request.form.get('password')

    if not all([email, password]):
        return render_template('index.html', message="⚠️ Please fill all fields")

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        fullname, username, stored_email, stored_password, stored_type = row
        if email == stored_email and password == stored_password:
            print(f"✅ Login successful: {fullname}")
            if stored_type == 'farmer':
                return redirect(url_for('dashboard', name=fullname, Type=stored_type))
            elif stored_type == 'vendor':
                return redirect(url_for('dashboardV', name=fullname, Type=stored_type))
            else:
                return render_template('index.html', message="⚠️ Please complete registration (choose your role).")

    print("❌ Invalid login attempt")
    return render_template('index.html', message="❌ Invalid Email Or Password Try Again Or Register New Account.")


# ---------------- DASHBOARDS ----------------
@app.route('/dashboard')
def dashboard():
    name = request.args.get('name', 'User')
    Type = request.args.get('Type', 'User')
    return render_template('dashboard.html', name=name, Type=Type)

@app.route('/dashboardV')
def dashboardV():
    name = request.args.get('name', 'User')
    Type = request.args.get('Type', 'User')
    return render_template('dashboardV.html', name=name, Type=Type)

# ---------------- ADD LAND (Vendor) ----------------
@app.route('/addLand', methods=['GET', 'POST'])
def addLand():
    if request.method == 'POST':
        location = request.form.get('location')
        area = request.form.get('area')

        if not all([location, area]):
            return render_template('addLand.html', message="⚠️ Please fill all fields")

        # Excel setup (vendor lands)
        LAND_FILE = os.path.join(BASE_DIR, "vendor_lands.xlsx")
        if not os.path.exists(LAND_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Lands"
            ws.append(["Location", "Area (acres)"])
            wb.save(LAND_FILE)

        wb = load_workbook(LAND_FILE)
        ws = wb.active
        ws.append([location, area])
        wb.save(LAND_FILE)

        print(f"✅ Land Added: {location} | {area} acres")
        return redirect(url_for('myLands'))

    return render_template('addLand.html')


# ---------------- MY LANDS (Vendor) ----------------
@app.route('/myLands')
def myLands():
    LAND_FILE = os.path.join(BASE_DIR, "vendor_lands.xlsx")

    lands = []
    if os.path.exists(LAND_FILE):
        wb = load_workbook(LAND_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            lands.append({"location": row[0], "area": row[1]})

    return render_template('myLands.html', lands=lands)
# Add near the top, after EXCEL_FILE
CONNECTION_FILE = os.path.join(BASE_DIR, "connections.xlsx")

def initialize_connections():
    if not os.path.exists(CONNECTION_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Connections"
        ws.append(["Vendor", "Farmer", "Status"])
        wb.save(CONNECTION_FILE)

@app.route('/connectFarmers')
def connectFarmers():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    farmers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fullname, username, email, password, user_type = row
        if user_type == "farmer":
            farmers.append(fullname)

    return render_template('connectFarmers.html', farmers=farmers)

@app.route('/sendRequest', methods=['POST'])
def sendRequest():
    # example logic — modify as per your app
    farmer_id = request.form.get('farmer_id')
    current_user = session.get('user_id')
    
    # (Optional) Save a connection request in database
    # db.execute("INSERT INTO connections (from_user, to_farmer) VALUES (?, ?)", (current_user, farmer_id))
    # db.commit()
    
    flash("Request sent successfully!", "success")
    return redirect(url_for('connectFarmers'))

# ---------------- START APP ----------------
if __name__ == "__main__":
    initialize_excel()
    initialize_connections()
    print("🚀 Flask app running at: http://127.0.0.1:5000/")
    app.run(debug=True)
